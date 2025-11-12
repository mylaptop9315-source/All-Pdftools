# main.py - Updated with Text to PDF + Render Webhook Ready

import os
import io
import asyncio
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
from PIL import Image
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont  # For custom fonts (optional Hindi)
from pdf2docx import Converter
import openpyxl
from weasyprint import HTML
import requests
import tempfile
import zipfile
import qrcode
from datetime import datetime

# Render pe webhook ke liye (polling ki jagah)
import threading
import time
from telegram.request import HTTPXRequest

# Apna Token yahan daalo (Render pe env var mein daalenge)
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")

# Webhook URL (Render deploy ke baad set karo)
WEBHOOK_URL = os.getenv("WEBHOOK_URL", None)

# Temp folder
if not os.path.exists("temp"):
    os.makedirs("temp")

# === /start ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("Photo → PDF", callback_data="photo_pdf")],
        [InlineKeyboardButton("PDF → Word", callback_data="pdf_word")],
        [InlineKeyboardButton("PDF → Excel", callback_data="pdf_excel")],
        [InlineKeyboardButton("Web → PDF", callback_data="web_pdf")],
        [InlineKeyboardButton("Rename File", callback_data="rename")],
        [InlineKeyboardButton("More Tools", callback_data="more")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "Namaste! Main **Free Converter Bot** hu.\n"
        "Neeche se tool chuno ya direct file bhejo!",
        reply_markup=reply_markup, parse_mode='Markdown'
    )

# === More Tools Menu (Text to PDF Added!) ===
async def more_tools(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("QR Code Banao", callback_data="qr")],
        [InlineKeyboardButton("Text → Speech", callback_data="tts")],
        [InlineKeyboardButton("Text → PDF", callback_data="text_pdf")],  # NEW!
        [InlineKeyboardButton("Compress Images", callback_data="compress")],
        [InlineKeyboardButton("ZIP Banao", callback_data="zip")],
        [InlineKeyboardButton("Merge PDFs", callback_data="merge_pdf")],
        [InlineKeyboardButton("Back", callback_data="back")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    query = update.callback_query
    await query.edit_message_text("Aur Tools (Text to PDF Added!):", reply_markup=reply_markup)

# === Button Handler (Text PDF Added) ===
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "photo_pdf":
        await query.edit_message_text("Photo bhejo → Main PDF bana dunga!")
        context.user_data['mode'] = 'photo_pdf'
    elif query.data == "pdf_word":
        await query.edit_message_text("PDF bhejo → Word (.docx) milega!")
        context.user_data['mode'] = 'pdf_word'
    elif query.data == "pdf_excel":
        await query.edit_message_text("PDF bhejo → Excel (.xlsx) milega!")
        context.user_data['mode'] = 'pdf_excel'
    elif query.data == "web_pdf":
        await query.edit_message_text("URL bhejo (http/https) → Webpage ka PDF banega!")
        context.user_data['mode'] = 'web_pdf'
    elif query.data == "rename":
        await query.edit_message_text("File bhejo aur reply mein likho:\n`rename newname.pdf`")
        context.user_data['mode'] = 'rename'
    elif query.data == "qr":
        await query.edit_message_text("Text bhejo → Main QR Code bana dunga!")
        context.user_data['mode'] = 'qr'
    elif query.data == "tts":
        await query.edit_message_text("Text bhejo → Main bolke audio dunga! (Hindi/English)")
        context.user_data['mode'] = 'tts'
    elif query.data == "text_pdf":  # NEW!
        await query.edit_message_text("Text bhejo (max 5000 chars) → Main PDF bana dunga! (Title: 'My Document')")
        context.user_data['mode'] = 'text_pdf'
    elif query.data == "compress":
        await query.edit_message_text("Photo bhejo → Chhota size wala milega!")
        context.user_data['mode'] = 'compress'
    elif query.data == "zip":
        await query.edit_message_text("1+ files bhejo → Main ZIP bana dunga! (/zipnow se finish)")
        context.user_data['mode'] = 'zip'
        context.user_data['zip_files'] = []
    elif query.data == "merge_pdf":
        await query.edit_message_text("2+ PDFs bhejo → Main ek PDF bana dunga! (/merge se finish)")
        context.user_data['mode'] = 'merge_pdf'
        context.user_data['pdfs'] = []
    elif query.data == "more":
        await more_tools(update, context)
    elif query.data == "back":
        await start(update, context)

# === Photo Handler (No Change) ===
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    mode = context.user_data.get('mode', '')
    
    if mode == 'photo_pdf' or not mode:
        photo = await update.message.photo[-1].get_file()
        photo_bytes = await photo.download_as_bytearray()
        img = Image.open(io.BytesIO(photo_bytes)).convert("RGB")

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        img = img.resize((int(width*0.9), int(height*0.9)), Image.Resampling.LANCZOS)
        img_io = io.BytesIO()
        img.save(img_io, 'JPEG')
        img_io.seek(0)
        c.drawImage(img_io, 30, height - img.height - 30, width=img.width, height=img.height)
        c.showPage()
        c.save()
        buffer.seek(0)

        await update.message.reply_document(buffer, filename="photo.pdf", caption="Photo → PDF Ready!")
        context.user_data['mode'] = None

    elif mode == 'compress':
        photo = await update.message.photo[-1].get_file()
        photo_bytes = await photo.download_as_bytearray()
        img = Image.open(io.BytesIO(photo_bytes)).convert("RGB")
        buffer = io.BytesOS()
        img.save(buffer, format='JPEG', quality=30)  # 70% chhota
        buffer.seek(0)
        await update.message.reply_document(buffer, filename="compressed.jpg", caption="Compressed Image!")
        context.user_data['mode'] = None

# === Document Handler (No Change) ===
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    mode = context.user_data.get('mode', '')

    if mode == 'pdf_word' and doc.mime_type == 'application/pdf':
        file_bytes = await doc.get_file().download_as_bytearray()
        with open("temp/temp.pdf", "wb") as f:
            f.write(file_bytes)
        cv = Converter("temp/temp.pdf")
        output = io.BytesIO()
        cv.convert(output)
        cv.close()
        output.seek(0)
        await update.message.reply_document(output, filename="converted.docx")
        os.remove("temp/temp.pdf")
        context.user_data['mode'] = None

    elif mode == 'pdf_excel' and doc.mime_type == 'application/pdf':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "PDF to Excel (Text Only)"
        ws['A2'] = "Advanced ke liye table wala PDF bhejo (future update)"
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        await update.message.reply_document(buffer, filename="output.xlsx")
        context.user_data['mode'] = None

    elif mode == 'zip':
        file_bytes = await doc.get_file().download_as_bytearray()
        context.user_data['zip_files'].append((doc.file_name, file_bytes))
        await update.message.reply_text(f"File added: {doc.file_name}\nTotal: {len(context.user_data['zip_files'])}\n /zipnow se ZIP banao!")

    elif mode == 'merge_pdf' and doc.mime_type == 'application/pdf':
        file_bytes = await doc.get_file().download_as_bytearray()
        context.user_data['pdfs'].append(file_bytes)
        await update.message.reply_text(f"PDF added! Total: {len(context.user_data['pdfs'])}\n /merge se merge karo!")

    elif mode == 'rename' and update.message.reply_to_message:
        new_name = update.message.text.strip()
        if new_name.startswith("rename "):
            new_name = new_name[7:].strip()
            if not new_name:
                new_name = "renamed_file"
            file = update.message.reply_to_message.document
            file_bytes = await file.get_file().download_as_bytearray()
            buffer = io.BytesIO(file_bytes)
            await update.message.reply_document(buffer, filename=new_name)
        else:
            await update.message.reply_text("Reply mein likho: `rename newfile.pdf`")

# === Text Handler (Text to PDF Added!) ===
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    mode = context.user_data.get('mode', '')

    if mode == 'web_pdf' and text.startswith("http"):
        try:
            html = requests.get(text, timeout=10).text
            pdf_buffer = io.BytesIO()
            HTML(string=html, base_url=text).write_pdf(pdf_buffer)
            pdf_buffer.seek(0)
            await update.message.reply_document(pdf_buffer, filename="webpage.pdf")
            context.user_data['mode'] = None
        except Exception as e:
            await update.message.reply_text(f"URL error: {str(e)}")

    elif mode == 'qr':
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(text)
        qr.make(fit=True)
        img = qr.make_image(fill='black', back_color='white')
        buffer = io.BytesIO()
        img.save(buffer, 'PNG')
        buffer.seek(0)
        await update.message.reply_photo(buffer, caption="QR Code Ready!")
        context.user_data['mode'] = None

    elif mode == 'tts':
        try:
            from gtts import gTTS
            lang = 'hi' if any(c in 'अआइईउऊऋएऐओऔ' for c in text) else 'en'
            tts = gTTS(text, lang=lang)
            buffer = io.BytesIO()
            tts.write_to_fp(buffer)
            buffer.seek(0)
            await update.message.reply_audio(buffer, filename="voice.mp3")
            context.user_data['mode'] = None
        except ImportError:
            await update.message.reply_text("TTS ke liye `pip install gtts` karo (requirements.txt mein hai).")

    elif mode == 'text_pdf':  # NEW! Text to PDF
        if len(text) > 5000:
            await update.message.reply_text("Text zyada lamba hai! 5000 chars tak rakho.")
            return
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        # Title
        c.setFont("Helvetica-Bold", 16)
        c.drawString(30, height - 50, "My Document")  # Title change kar sakte ho
        # Text (wrap karne ke liye simple loop)
        y = height - 100
        lines = text.split('\n')
        for line in lines:
            if y < 50:  # New page
                c.showPage()
                y = height - 50
            c.setFont("Helvetica", 10)
            if len(line) > 80:  # Wrap long lines
                words = line.split()
                current_line = ""
                for word in words:
                    if c.stringWidth(current_line + word, "Helvetica", 10) < width - 60:
                        current_line += word + " "
                    else:
                        c.drawString(30, y, current_line.strip())
                        y -= 15
                        current_line = word + " "
                        if y < 50:
                            c.showPage()
                            y = height - 50
                if current_line:
                    c.drawString(30, y, current_line.strip())
                    y -= 15
            else:
                c.drawString(30, y, line)
                y -= 15
        c.save()
        buffer.seek(0)
        await update.message.reply_document(buffer, filename="text_document.pdf", caption="Text → PDF Ready! 📄")
        context.user_data['mode'] = None

# === ZIP & Merge (No Change) ===
async def finish_zip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get('zip_files'):
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zf:
            for name, data in context.user_data['zip_files']:
                zf.writestr(name, data)
        zip_buffer.seek(0)
        await update.message.reply_document(zip_buffer, filename="files.zip")
        context.user_data['zip_files'] = []
        context.user_data['mode'] = None
    else:
        await update.message.reply_text("Koi file nahi hai ZIP ke liye.")

async def finish_merge(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.user_data.get('pdfs', [])) >= 2:
        try:
            from PyPDF2 import PdfMerger
            merger = PdfMerger()
            for pdf_data in context.user_data['pdfs']:
                merger.append(io.BytesIO(pdf_data))
            output = io.BytesIO()
            merger.write(output)
            merger.close()
            output.seek(0)
            await update.message.reply_document(output, filename="merged.pdf")
            context.user_data['pdfs'] = []
            context.user_data['mode'] = None
        except ImportError:
            await update.message.reply_text("PyPDF2 install karo (requirements.txt mein hai).")
    else:
        await update.message.reply_text("Kam se kam 2 PDFs bhejo!")

# === Webhook Setup for Render ===
async def set_webhook(app):
    if WEBHOOK_URL:
        await app.bot.set_webhook(url=f"{WEBHOOK_URL}/webhook")
        print("Webhook set!")

# === Main (Webhook Mode for Render) ===
def main():
    request = HTTPXRequest(connect_timeout=10.0, read_timeout=10.0)
    app = Application.builder().token(BOT_TOKEN).request(request).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(CommandHandler("zipnow", finish_zip))
    app.add_handler(CommandHandler("merge", finish_merge))

    if WEBHOOK_URL:
        # Webhook mode
        app.run_webhook(
            listen="0.0.0.0",
            port=int(os.environ.get("PORT", 8443)),
            url_path="/webhook",
            webhook_url=f"{WEBHOOK_URL}/webhook"
        )
    else:
        # Local testing ke liye polling
        print("Local mode: Polling...")
        app.run_polling()

if __name__ == "__main__":
    main()
